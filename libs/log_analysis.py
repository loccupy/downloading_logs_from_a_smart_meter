import datetime
from datetime import timedelta
from time import sleep

from dateutil.relativedelta import relativedelta
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

from libs.Utils import is_valid_date_for_anal, parse_log_name

pink_fill = PatternFill(start_color='FFCC99FF',
                        end_color='FFCC99FF',
                        fill_type='solid')


# считать excel gegefe
def read_excel(path_to_file):
    try:
        sheets = load_workbook(path_to_file)

        # analysis_correct_date_ref(sheets, 'Журнал напряжения')
        # time_ordering_analysis_ref(sheets, 'Журнал коммуникационных событий')
        # ipu_working_hours_ref(sheets, 'Журнал напряжения')

        sheets.save(path_to_file)

    except Exception as e:
        raise


# анализирует "Время фиксации записи" на предмет невалидных дат, FFF-ок
def analysis_correct_date_ref(sheets, log_name):
    try:
        sheet = sheets[log_name]
        column = None
        column_names = [cell for cell in sheet[1]]
        # Получаем номер столбца по названию
        for i in column_names:
            if i.value == 'Время фиксации записи':
                column = i.column

        # Проверяем все значения столбца и, в случае совпадения условия, перекрашиваем ячейку
        for row in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
            # Получаем ячейку
            cell = row[0]
            # Проверяем условие
            if not is_valid_date_for_anal(str(cell.value)):
                cell.fill = pink_fill
                sleep(0.1)
                print(f"Невалидная дата в строчке {cell.row} в '{parse_log_name(log_name)}'")
    except Exception as e:
        raise f"Ошибка {e} при анализе 'Время фиксации записи' на последовательность в {parse_log_name(log_name)}"


# проверяет очередность валидных дат в столбце "Время фиксации записи"
def time_ordering_analysis_ref(sheets, log_name):
    format_dt = "%d.%m.%y %H:%M:%S"
    try:
        sheet = sheets[log_name]
        column = None
        column_names = [cell for cell in sheet[1]]
        # Получаем номер столбца по названию
        for i in column_names:
            if i.value == 'Время фиксации записи':
                column = i.column
                break

        # Проверяем все значения столбца и, в случае совпадения условия, перекрашиваем ячейку
        for i, row in enumerate(sheet.iter_rows(min_row=2, min_col=column, max_col=column)):
            # Получаем и сохраняем первую ячейку + пропускаем ее для обработки
            if i == 0:
                cell_0 = row[0]
                continue

            # Получаем ячейку
            cell_1 = row[0]

            if is_valid_date_for_anal(str(cell_0.value)) and is_valid_date_for_anal(str(cell_1.value)):
                # Проверяем условие
                if ((datetime.datetime.strptime(str(cell_0.value), format_dt) -
                     datetime.datetime.strptime(str(cell_1.value), format_dt) > timedelta())):
                    cell_0.fill = pink_fill
                    sleep(0.1)
                    cell_1.fill = pink_fill
                    sleep(0.1)

                    print(
                        f"Некорректная последовательность фиксации записи в строчках {cell_0.row}-{cell_1.row} в"
                        f" '{parse_log_name(log_name)}'")
            cell_0 = row[0]
    except Exception as e:
        raise f"Ошибка {e} при анализе 'Время фиксации записи' на валидность в {parse_log_name(log_name)}"


# Проверяет последовательность значений в столбце "Время работы ИПУ"
def ipu_working_hours_ref(sheets, log_name):
    try:
        sheet = sheets[log_name]
        column = None
        column_names = [cell for cell in sheet[1]]
        # Получаем номер столбца по названию
        for i in column_names:
            if i.value == 'Время работы ПУ':
                column = i.column
                break

        # Проверяем все значения столбца и, в случае совпадения условия, перекрашиваем ячейку
        for i, row in enumerate(sheet.iter_rows(min_row=2, min_col=column, max_col=column)):
            # Получаем и сохраняем первую ячейку + пропускаем ее для обработки
            if i == 0:
                cell_0 = row[0]
                continue

            # Получаем ячейку
            cell_1 = row[0]
            if cell_1.value is None:
                cell_1.fill = pink_fill
                sleep(0.1)
                print(f"Некорректный тип данных 'Время работы ПУ' в ячейке {cell_1.row} в"
                      f" '{parse_log_name(log_name)}'")
                continue
            else:
                res = int(cell_1.value) - int(cell_0.value)

                # Проверяем условие
                if res < 0:
                    cell_0.fill = pink_fill
                    sleep(0.1)
                    cell_1.fill = pink_fill
                    sleep(0.1)

                    print(f"Некорректная последовательность 'Время работы ПУ' в строчках {cell_0.row}-{cell_1.row} в"
                          f" '{log_name.replace(" ", "е ", 1)}'")

                cell_0 = row[0]

    except Exception as e:
        raise f"Ошибка {e} при анализе 'Время работы ИПУ' в {parse_log_name(log_name)}"


# проверяет очередность валидных дат в столбце "Время фиксации записи" в суточном профиле
def time_ordering_analysis_for_daily_profile(sheets, log_name):
    format_dt = "%d.%m.%y %H:%M:%S"
    try:
        sheet = sheets[log_name]
        column = None
        column_names = [cell for cell in sheet[1]]
        # Получаем номер столбца по названию
        for i in column_names:
            if i.value == 'Время фиксации записи':
                column = i.column
                break

        # Проверяем все значения столбца и, в случае совпадения условия, перекрашиваем ячейку
        for i, row in enumerate(sheet.iter_rows(min_row=3, min_col=column, max_col=column)):
            # Получаем и сохраняем первую ячейку + пропускаем ее для обработки
            if i == 0:
                cell_0 = row[0]
                continue

            # Получаем ячейку
            cell_1 = row[0]

            if is_valid_date_for_anal(str(cell_0.value)) and is_valid_date_for_anal(str(cell_1.value)):
                delta = (datetime.datetime.strptime(str(cell_1.value), format_dt) -
                         datetime.datetime.strptime(str(cell_0.value), format_dt))
                # Проверяем условие
                if delta != timedelta(days=1):
                    cell_0.fill = pink_fill
                    sleep(0.1)
                    cell_1.fill = pink_fill
                    sleep(0.1)

                    print(
                        f"Некорректная последовательность фиксации записи в строчках {cell_0.row}-{cell_1.row} в"
                        f" '{parse_log_name(log_name)}'")
            cell_0 = row[0]
    except Exception as e:
        raise f"Ошибка {e} при анализе 'Время фиксации записи' в {parse_log_name(log_name)}"


# проверяет очередность валидных дат в столбце "Время фиксации записи" в месячном профиле
def time_ordering_analysis_for_month_profile(sheets, log_name):
    format_dt = "%d.%m.%y %H:%M:%S"
    try:
        sheet = sheets[log_name]
        column = None
        column_names = [cell for cell in sheet[1]]
        # Получаем номер столбца по названию
        for i in column_names:
            if i.value == 'Время фиксации записи':
                column = i.column
                break

        # Проверяем все значения столбца и, в случае совпадения условия, перекрашиваем ячейку
        for i, row in enumerate(sheet.iter_rows(min_row=3, min_col=column, max_col=column)):
            # Получаем и сохраняем первую ячейку + пропускаем ее для обработки
            if i == 0:
                cell_0 = row[0]
                continue

            # Получаем ячейку
            cell_1 = row[0]

            if is_valid_date_for_anal(str(cell_0.value)) and is_valid_date_for_anal(str(cell_1.value)):
                date_1 = datetime.datetime.strptime(str(cell_0.value), format_dt)
                date_2 = datetime.datetime.strptime(str(cell_1.value), format_dt)
                delta = relativedelta(date_2, date_1)
                # Проверяем условие
                if not (delta.months == 1 and delta.days == 0):
                    cell_0.fill = pink_fill
                    sleep(0.1)
                    cell_1.fill = pink_fill
                    sleep(0.1)

                    print(
                        f"Некорректная последовательность фиксации записи в строчках {cell_0.row}-{cell_1.row} в"
                        f" '{parse_log_name(log_name)}'")
            cell_0 = row[0]
    except Exception as e:
        raise f"Ошибка {e} при анализе 'Время фиксации записи' в {parse_log_name(log_name)}"


#Функция проверки ошибок самодиагностики
def checkForSelfDiagnostics(sheets, log_name):
    #Словарь ошибок самодиагностики
    errorsList = ["2, Измерительный блок - ошибка",
                  "4, Вычислительный блок - ошибка",
                  "5, Часы реального времени - ошибка",
                  "7, Блок питания - ошибка",
                  "9, Дисплей - ошибка",
                  "11, Блок памяти - ошибка",
                  "13, Блок памяти программ - ошибка",
                  "15, Система тактирования ядра - ошибка",
                  "17, Система тактирования часов - ошибка",
                  "129, Аппаратный сброс часов реального времени",
                  "132, Сброс микроконтроллера сторожевым таймером"]

    try:
        sheet = sheets[log_name]
        column = None
        column_names = [cell for cell in sheet[1]]
        # Получаем номер столбца по названию
        for i in column_names:
            if i.value == "Событие":
                column = i.column
                break

        # Проверяем все значения столбца и, в случае совпадения условия, перекрашиваем ячейку
        for row in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
            if row[0].value in errorsList:
                row[0].fill = pink_fill
                sleep(0.1)
                print(
                    f"Код ошибки {row[0].value.split(',')[0]} в строчке {row[0].row} в"
                    f" '{parse_log_name(log_name)}'")

    except Exception as e:
        raise f"Ошибка {e} при анализе кодов ошибок в {parse_log_name(log_name)}"


# проверяет повторные включения или выключения в журнале вкл\выкл
def checking_for_repeated_on_or_offs(sheets, log_name):
    format_dt = "%d.%m.%y %H:%M:%S"
    try:
        sheet = sheets[log_name]
        column = None
        column_names = [cell for cell in sheet[1]]
        # Получаем номер столбца по названию
        for i in column_names:
            if i.value == 'Событие':
                column = i.column
                break

        # Проверяем все значения столбца и, в случае совпадения условия, перекрашиваем ячейку
        for i, row in enumerate(sheet.iter_rows(min_row=2, min_col=column, max_col=column)):
            # Получаем и сохраняем первую ячейку + пропускаем ее для обработки
            if i == 0:
                cell_0 = row[0]
                continue

            # Получаем ячейку
            cell_1 = row[0]

            if str(cell_0.value) == str(cell_1.value):
                # Проверяем условие

                cell_0.fill = pink_fill
                sleep(0.1)
                cell_1.fill = pink_fill
                sleep(0.1)

                print(
                    f"Повторяется код события в строчках {cell_0.row}-{cell_1.row} в"
                    f" '{parse_log_name(log_name)}'")
            cell_0 = row[0]
    except Exception as e:
        raise f"Ошибка {e} при анализе повторных кодов событий в {parse_log_name(log_name)}"


# проверяет очередность валидных дат в столбце "Время фиксации записи" в срезе мгновенных значений
def time_ordering_analysis_for_artur_profile(sheets, log_name):
    format_dt = "%d.%m.%y %H:%M:%S"
    try:
        sheet = sheets[log_name]
        column = None
        column_names = [cell for cell in sheet[1]]
        # Получаем номер столбца по названию
        for i in column_names:
            if i.value == 'Время фиксации записи':
                column = i.column
                break

        # Проверяем все значения столбца и, в случае совпадения условия, перекрашиваем ячейку
        for i, row in enumerate(sheet.iter_rows(min_row=2, min_col=column, max_col=column)):
            # Получаем и сохраняем первую ячейку + пропускаем ее для обработки
            if i == 0:
                cell_0 = row[0]
                continue

            # Получаем ячейку
            cell_1 = row[0]

            if is_valid_date_for_anal(str(cell_0.value)) and is_valid_date_for_anal(str(cell_1.value)):
                delta = (datetime.datetime.strptime(str(cell_1.value), format_dt) -
                         datetime.datetime.strptime(str(cell_0.value), format_dt))
                # Проверяем условие
                if delta != timedelta(minutes=30):
                    cell_0.fill = pink_fill
                    sleep(0.1)
                    cell_1.fill = pink_fill
                    sleep(0.1)

                    print(
                        f"Некорректная последовательность фиксации записи в строчках {cell_0.row}-{cell_1.row} в"
                        f" '{parse_log_name(log_name)}'")
                    add_to_global_list(
                        f"Некорректная последовательность фиксации записи в строчках {cell_0.row}-{cell_1.row} в"
                        f" '{parse_log_name(log_name)}'")
            cell_0 = row[0]
    except Exception as e:
        raise f"Ошибка {e} при анализе 'Время фиксации записи' в {parse_log_name(log_name)}"