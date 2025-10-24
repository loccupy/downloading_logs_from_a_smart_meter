from openpyxl.reader.excel import load_workbook

from libs.log_analysis import ipu_working_hours_ref, analysis_correct_date_ref, time_ordering_analysis_ref, \
    time_ordering_analysis_for_daily_profile, time_ordering_analysis_for_month_profile, checkForSelfDiagnostics, \
    checking_for_repeated_on_or_offs, time_ordering_analysis_for_artur_profile


def current_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал токов'
        print("Анализ журнала токов...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def self_diagnosis_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал самодиагностики'
        print("Анализ журнала самодиагностики...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)
        checkForSelfDiagnostics(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def network_quality_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал качества сети'
        print("Анализ журнала качества сети...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def voltage_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал напряжения'
        print("Анализ журнала напряжения...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def communication_events_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал коммуникационных событий'
        print("Анализ журнала коммуникационных событий...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def access_control_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал контроля доступа'
        print("Анализ журнала контроля доступа...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def data_correction_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал коррекции данных'
        print("Анализ журнала коррекции данных...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)
        raise


def time_correction_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал коррекции времени'
        print("Анализ журнала коррекции времени...")

        analysis_correct_date_ref(sheets, log_name)
        # time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def battery_charge_status_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал состояния заряда батареи'
        print("Анализ журнала состояния заряда батареи...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def power_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал мощности'
        print("Анализ журнала мощности...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Превышение тангенса
def tangent_excess_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал превышения тангенса'
        print("Анализ журнала превышения тангенса...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Выход тангенса
def tangent_output_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал Выход тангенса'
        print("Анализ журнала Выход тангенса...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def network_quality_for_period_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал качества сети за период'
        print("Анализ журнала качества сети за период...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


def on_and_off_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал включений и выключений'
        print("Анализ журнала включений и выключений...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)
        checking_for_repeated_on_or_offs(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Внешних воздействий
def external_influences_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал внешних воздействий'
        print("Анализ журнала внешних воздействий...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Дискретки
def sampling_status_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Журнал состояний дискреток'
        print("Анализ журнала состояний дискреток...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Суточный профиль
def daily_profile_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Суточный профиль'
        print("Анализ журнала Суточный профиль...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        time_ordering_analysis_for_daily_profile(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Месячный профиль
def month_profile_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Месячный профиль'
        print("Анализ журнала Месячный профиль...")

        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        time_ordering_analysis_for_month_profile(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Профиль энергии за инт. 1
def energy_profile_for_1_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Профиль энергии за инт. 1'
        print("Анализ журнала Профиль энергии за инт. 1...")
        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Профиль энергии за инт. 2
def energy_profile_for_2_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Профиль энергии за инт. 2'
        print("Анализ журнала Профиль энергии за инт. 2...")
        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)


# Срез мгновенных значений
def artur_profile_log_analysis(path_to_file):
    try:
        sheets = load_workbook(path_to_file)
        log_name = 'Срез мгновенных значений'
        print("Анализ журнала Срез мгновенных значений...")
        analysis_correct_date_ref(sheets, log_name)
        time_ordering_analysis_ref(sheets, log_name)
        ipu_working_hours_ref(sheets, log_name)
        time_ordering_analysis_for_artur_profile(sheets, log_name)

        sheets.save(path_to_file)

    except Exception as e:
        print(e)
