class RunRelativeRecord:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.workbook = self._load_workbook()

    def _load_workbook(self):
        from openpyxl import load_workbook
        return load_workbook(self.excel_file_path)

    def extract_time_data(self):
        """
        Считывает значения из столбцов:
        - "Время фиксации записи"
        - "Время работы ПУ"
        со всех листов таблицы.
        Возвращает список словарей с координатами и значениями ячеек.
        """
        data = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            time_fix_col = None
            pu_time_col = None

            # Поиск заголовков в первой строке
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value == "Время фиксации записи":
                    time_fix_col = col_idx
                elif cell.value == "Время работы ПУ":
                    pu_time_col = col_idx

            # Считывание данных по строкам
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_data = {
                    'sheet': sheet_name,
                    'row': row[0].row,
                    'time_fix': None,
                    'time_fix_cell': None,
                    'pu_time': None,
                    'pu_time_cell': None,
                }

                if time_fix_col and row[time_fix_col - 1].value is not None:
                    cell = row[time_fix_col - 1]
                    row_data['time_fix'] = cell.value
                    row_data['time_fix_cell'] = f"{cell.column_letter}{cell.row}"

                if pu_time_col and row[pu_time_col - 1].value is not None:
                    cell = row[pu_time_col - 1]
                    row_data['pu_time'] = cell.value
                    row_data['pu_time_cell'] = f"{cell.column_letter}{cell.row}"
                else:
                    continue

                data.append(row_data)

        return data

    def find_time_fix_matches(self, time_data, tolerance_minutes=0.1):
        """
        Ищет совпадения между 'time_fix' с одного листа и 'time_fix' с другого листа,
        если они находятся в пределах указанной погрешности (по умолчанию — 2 минуты).

        :param time_data: результат вызова extract_time_data()
        :param tolerance_minutes: допуск в минутах для сравнения времени
        :return: список совпадений с информацией о ячейках и разнице во времени
        """
        from datetime import timedelta

        matches = []
        tolerance = timedelta(minutes=tolerance_minutes)

        # Фильтруем записи, где есть time_fix
        fix_records = [record for record in time_data if
                       record['time_fix'] is not None and record['pu_time'] is not None]

        for i, record1 in enumerate(fix_records):
            for record2 in fix_records[i + 1:]:
                # Пропускаем сравнение записей с одного листа, если нужно сравнивать только между листами
                if record1['sheet'] == record2['sheet'] or record1['sheet'] == "Журнал самодиагностики":
                    continue

                t1, t2 = record1['time_fix'], record2['time_fix']

                # Преобразуем строки в datetime при необходимости
                if isinstance(t1, str):
                    try:
                        from dateutil import parser
                        t1 = parser.parse(t1)
                    except Exception:
                        continue
                if isinstance(t2, str):
                    try:
                        from dateutil import parser
                        t2 = parser.parse(t2)
                    except Exception:
                        continue

                # Проверяем разницу
                if abs(t1 - t2) <= tolerance:
                    matches.append({
                        'sheet_1': record1['sheet'],
                        'cell_1': record1['time_fix_cell'],
                        'value_1': t1,
                        'value_1_2': record1['pu_time'],
                        'sheet_2': record2['sheet'],
                        'cell_2': record2['time_fix_cell'],
                        'value_2': t2,
                        'value_2_2': record2['pu_time'],
                        'time_diff_1': abs(t1 - t2).seconds,
                        'time_diff_2': abs(record1['pu_time'] - record2['pu_time'])
                    })

        return matches

    def find_false_res(self, matches):
        from datetime import timedelta
        for match in matches:
            try:
                delta1 = match['time_diff_1']
                delta2 = match['time_diff_2']
                diff = abs(delta1-delta2)
                if diff > 3:
                    print("Несоответствие интервалов:", match)
                    print(f"Интервал 1: {delta1}, Интервал 2: {delta2}")
                    print(f"Разница: {diff}\n")
            except KeyError as e:
                print(f"Не хватает поля: {e}")
            except TypeError as e:
                print(f"Ошибка типа (не datetime?): {e}")
